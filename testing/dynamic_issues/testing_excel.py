import openpyxl

wb = openpyxl.load_workbook(file)
old_sheet = wb.get_sheet_by_name('Sheet1')
old_sheet.title = 'Sheet1.5'
max_row = old_sheet.get_highest_row()
max_col = old_sheet.get_highest_column()
wb.create_sheet(0, 'Sheet1')

new_sheet = wb.get_sheet_by_name('Sheet1')

# Do the header.
for col_num in range(0, max_col):
    new_sheet.cell(row=0, column=col_num).value = old_sheet.cell(row=0, column=col_num).value

# The row to be inserted. We're manually populating each cell.
new_sheet.cell(row=1, column=0).value = 'DUMMY'
new_sheet.cell(row=1, column=1).value = 'DUMMY'

# Now do the rest of it. Note the row offset.
for row_num in range(1, max_row):
    for col_num in range (0, max_col):
        new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

wb.save(file)
